import openpyxl

workbook = openpyxl.load_workbook("Reest Material 2023.xlsx")
sheet = workbook.active


def write_excel(number, content, date, where, reason, owner):
    number_propusk = None
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=6, values_only=True):
        if row[0] is not None:
            number_propusk = row[0]

    next_row_index = sheet.max_row + 1 if number_propusk is not None else 1
    sheet.cell(row=next_row_index, column=1, value=number)
    sheet.cell(row=next_row_index, column=2, value=content)
    sheet.cell(row=next_row_index, column=3, value=date)
    sheet.cell(row=next_row_index, column=4, value=where)
    sheet.cell(row=next_row_index, column=5, value=reason)
    sheet.cell(row=next_row_index, column=6, value=owner)
    workbook.save("Reest Material 2023.xlsx")


def last_value():
    last_cell_value = None
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=1, values_only=True):
        if row[0] is not None:
            last_cell_value = row[0]
    return int(last_cell_value)
